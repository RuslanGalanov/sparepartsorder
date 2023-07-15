import openpyxl

from typing import Tuple
from .models import Spare



def parse(file: str, range_rows: list[Tuple[int, int]]) -> list[Spare]:

    wb = openpyxl.load_workbook(file)
    sheet = wb.active

    spares = []
    stop = True

    for position in range_rows:

        for row in range(position[0], position[1] + 1, 2):
            
            amount = int(sheet.cell(row=row, column=2).value)
            name = sheet.cell(row=row, column=1).value.strip().capitalize()
            ref = sheet.cell(row=row+1, column=1).value or None

            if ref: 
                ref = ref.strip().upper()

            spare = Spare(ref=ref, name=name, year=file[0:4], amount=amount)

            for item in spares:
                if (spare.name != item.name) and (spare.ref != item.ref):
                    stop = True
                elif (spare.name == item.name) and (spare.ref == item.ref):
                    item.amount += spare.amount
                    stop = False
                    break

            if stop:
                spares.append(spare)
            else:
                continue

    return spares


def save_data(filename: str, data: list[tuple]):

    new_wb = openpyxl.Workbook()

    sheet = new_wb.active

    for i, result in enumerate(data):
        sheet.cell(row=i+1, column=1).value = result[0]
        sheet.cell(row=i+1, column=2).value = result[1]
        sheet.cell(row=i+1, column=3).value = str(result[2])
        sheet.cell(row=i+1, column=4).value = result[3]
        sheet.cell(row=i+1, column=5).value = result[4]
        sheet.cell(row=i+1, column=6).value = result[5]
        sheet.cell(row=i+1, column=8).value = ' '.join(result[6])


    new_wb.save(filename)